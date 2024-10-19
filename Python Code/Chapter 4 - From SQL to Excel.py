# =============================================================================
# Chapter 4 - From SQL to Excel.py
# =============================================================================

import os
import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# Establish a connection to the SQL Server
connection_string = ('Driver={ODBC Driver 17 for SQL Server};'
                    'Server=Enter_Your_Server_Name;' # Replace with your SQL Server name
                    'Database=Automate_Everything_With_Python;'
                    'UID=automation_user;'
                    'PWD=AutomateEverything2024!;'
                    'Trusted_Connection=no')

conn = pyodbc.connect(connection_string)

# Define the SQL query
sql_query = """
            SELECT 
                   p.ProviderID,
                   p.ProviderName,
                   p.Phone,
                   pa.ClinicID,
                   pa.PatientID,
                   pa.PatientName,
                   pa.Temperature,
                   pa.Pulse_Oximeter,
                   pa.DateTimeAssessment 
            FROM Automate_Everything_With_Python.dbo.PatientAssessments pa
            LEFT JOIN Automate_Everything_With_Python.dbo.Providers p ON pa.ProviderID = p.ProviderID"""

# Execute the query and fetch the results into a DataFrame
df_patient_readings = pd.read_sql(sql_query, conn)

# Close the connection
conn.close()

# Convert DateTimeAssessment to datetime format
df_patient_readings['DateTimeAssessment'] = pd.to_datetime(df_patient_readings['DateTimeAssessment'])

# Sort and select the last three records per patient
df_sorted = df_patient_readings.sort_values(by=['PatientID', 'DateTimeAssessment'])
df_last_three = df_sorted.groupby('PatientID').tail(3)
df_last_three.reset_index(drop=True, inplace=True)

# Define paths
desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop')
report_folder_path = os.path.join(desktop_path, 'Automate Everything With Python', 'Automated Reports')
os.makedirs(report_folder_path, exist_ok=True)

# Process each provider's data
for provider_id in df_last_three['ProviderID'].unique():
    provider_data = df_last_three[df_last_three['ProviderID'] == provider_id]
    provider_name = provider_data['ProviderName'].values[0]
    excel_file_name = f"{provider_name} - {provider_id} Morning Readings.xlsx"
    excel_file_path = os.path.join(report_folder_path, excel_file_name)
    provider_data.to_excel(excel_file_path, index=False)

    # Load workbook and worksheet
    wb = load_workbook(excel_file_path)
    ws = wb.active
    
    # Rename worksheet and turn off gridlines
    ws.title = "Data"
    ws.sheet_view.showGridLines = False
    
    # Create a table in the worksheet
    table_name = "DataTable"
    table = Table(displayName=table_name, ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium18", 
                           showFirstColumn=False, 
                           showLastColumn=False, 
                           showRowStripes=True, 
                           showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Adjust column widths
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    # Apply conditional formatting
    temp_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
    oxygen_fill = PatternFill(start_color='83F28F', end_color='83F28F', fill_type='solid')
    for row in range(2, ws.max_row + 1):
        temperature_cell = ws.cell(row=row, column=7)
        pulse_oximeter_cell = ws.cell(row=row, column=8)
        if temperature_cell.value is not None and temperature_cell.value >= 100:
            temperature_cell.fill = temp_fill
        if pulse_oximeter_cell.value is not None and pulse_oximeter_cell.value <= 95:
            pulse_oximeter_cell.fill = oxygen_fill

    # Save the workbook
    wb.save(excel_file_path)
